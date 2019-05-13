import bs4
import requests
from openpyxl import Workbook


link='https://www.snapdeal.com/products/computers-webcams?sort=plrty'
#link = input()
resultset = requests.get(link)


soup = bs4.BeautifulSoup(resultset.content,'html.parser')


wb=Workbook()
ws = wb.active
ws.cell(row=1,column=1,value='Product ID')
ws.cell(row=1,column=2,value='striked value')
ws.cell(row=1,column=3,value='Price')
#ws.cell(row=1,column=4,value='No. of times rated')


allprods= soup.find_all('div',class_='product-desc-rating')

for i in range(len(allprods)):
    ws.cell(row=i+2,column=1,value=allprods[i].a['pogid'])
    strk=allprods[i].a.find('span',class_="lfloat product-desc-price strike").text
    ws.cell(row=i+2,column=2,value=strk[3:])
    pprc=allprods[i].a.find('span',class_="lfloat product-price")
    ws.cell(row=i+2,column=3,value=pprc['data-price'])
    #x = allprods[i].a.find('p',class_="product-rating-count").text
    #rating =allprods[i].a.find('p',class_="product-rating-count").text
    #ws.cell(row=i+2,column=4,value=rating[1:-1])
    
wb.save('snapscrap.xlsx')

#firstone= allprods[0]

#firstone.a['pogid']
#firstone.a
#firstone.a.find('span',class_="lfloat product-desc-price strike")
#firstone.a.find('span',class_="lfloat product-price")
#pprc=firstone.a.find('span',class_="lfloat product-price")
#pprc['data-price']
#firstone.a.find('p',class_="product-rating-count").text
